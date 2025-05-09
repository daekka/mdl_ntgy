import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from pydantic import BaseModel
from openai import AzureOpenAI
import httpx
import json
from openpyxl.styles import Alignment
import ast

st.set_page_config(
    page_title="Puntuaciones M+ LLM",
    page_icon="🗯️",
    layout="wide",
)

# CSS personalizado para cambiar ancho del sidebar
st.markdown("""
    <style>
        section[data-testid="stSidebar"] {
            width: 275px !important;
        }
        section[data-testid="stSidebar"] > div:first-child {
            width: 275px !important;
        }
    </style>
""", unsafe_allow_html=True)


# Se reemplaza la carga directa del archivo por un file uploader
system_instructions = ""


def cargar_config_azure(archivo_config):
    try:
        contenido = archivo_config.read().decode('utf-8')
        config_dict = json.loads(contenido)
        
        # Verificar claves necesarias
        claves_requeridas = {"deployment_name", "api_key", "azure_endpoint", "api_version"}
        if not claves_requeridas.issubset(config_dict):
            st.error("Faltan una o más claves requeridas en la configuración.")
            return None

        return config_dict

    except json.JSONDecodeError:
        st.error("El archivo no tiene un formato JSON válido.")
        return None
    except Exception as e:
        st.error(f"Error al cargar la configuración: {str(e)}")
        return None


# Función para inicializar el cliente OpenAI
def inicializar_cliente_openai(config):
    httpx_client = httpx.Client(http2=True, verify=False)
    client = AzureOpenAI(
        api_key="cfe23"+config["api_key"],  
        azure_endpoint=config["azure_endpoint"],
        api_version=config["api_version"],
        default_headers={"Content-Type": "application/json; charset=utf-8"},
        http_client=httpx_client
    )
    return client

# Define tu modelo Pydantic
class puntuaciones(BaseModel):
    Severidad: str
    Probabilidad: str
    Ámbito: str

def LLM_Consulta(client, system_prompt = "", descripcion =""):
    try:
        # Realiza la solicitud al modelo desplegado
        completion = client.chat.completions.create(
            model="gpt-4o",  # Este es el nombre del *deployment*, no el modelo base. Usa el que configuraste en Azure.
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": descripcion},
            ],
            temperature=0.3 ,
            #response_format=puntuaciones,
            response_format={"type": "json_object"},
        )

        # Accede al contenido de la respuesta
        event = completion.choices[0].message.content
        #st.write(event)
        return(event)
    except Exception as e:
        st.error(f"Error en la consulta al LLM: {str(e)}")
        # Devuelve un JSON con valores nulos en caso de error
        return json.dumps({
            "SEVERIDAD": None,
            "PROBABILIDAD": None,
            "AMBITO": None,
            "PREGUNTAS": None,
            "ANALISIS": None,
            "RIESGOS": None,
            "RECOMENDACIONES": None
        })



st.title("Puntuaciones M+ con LLM🗯️")
with st.expander("Configuración 📋", expanded=True):
    # Crear tres columnas para los file uploaders
    col1, col2, col3 = st.columns(3)

    # Columna 1: Configuración de Azure
    with col1:
        st.subheader("Configuración AZURE")
        config_file = st.file_uploader("Sube tu archivo de configuración de Azure", type=["txt"])
        
    # Columna 2: Prompt del sistema
    with col2:
        st.subheader("Prompt")
        prompt_file = st.file_uploader("Sube tu archivo con el prompt del sistema", type=["txt", "md"])

    # Columna 3: Archivo Excel
    with col3:
        st.subheader("Excel Puntuaciones")
        uploaded_file = st.file_uploader("Sube tu archivo Excel", type=["xlsx"])

    # Cargar el prompt del sistema desde el archivo subido
    if prompt_file is not None:
        try:
            system_instructions = prompt_file.read().decode('utf-8')
            #st.success("Prompt del sistema cargado correctamente.")
        except Exception as e:
            st.error(f"Error al cargar el prompt del sistema: {str(e)}")

client = None

if config_file is not None:
    config = cargar_config_azure(config_file)
    if config:
        AZURE_CONFIG = config
        #st.success("Configuración cargada correctamente.")
        client = inicializar_cliente_openai(AZURE_CONFIG)
    else:
        st.error("No se pudo cargar la configuración.")

if uploaded_file is not None and client is not None:
    # Verifica que se haya cargado un prompt
    if not system_instructions:
        st.warning("Por favor, carga el archivo con el prompt del sistema antes de continuar.")
    else:
        in_memory_file = BytesIO(uploaded_file.read())
        
        st.divider()
        with st.expander("System Prompt 📋", expanded=False):
            st.markdown(system_instructions)
        st.divider()
        # Cargar libro con openpyxl
        wb = load_workbook(filename=in_memory_file)
        hojas = wb.sheetnames
        hoja_seleccionada = st.selectbox("Selecciona la hoja a tratar", hojas, index=hojas.index("Paralizaciones"))

        if hoja_seleccionada:
            ws = wb[hoja_seleccionada]
            data = list(ws.values)

            cabecera_index = None
            for i, fila in enumerate(data):
                if fila and "Severidad" in fila and "Probabilidad" in fila and "Ámbito" in fila:
                    cabecera_index = i
                    columnas = list(fila)

                    # Buscar la(s) columna(s) que contienen "Descripción"
                    columnas_descripcion = [col for col in columnas if "Descripción" in str(col)]
                    
                    if columnas_descripcion:
                        print(f"Columna(s) con 'Descripción' encontrada(s): {columnas_descripcion}")
                    else:
                        print("No se encontró ninguna columna con 'Descripción'")



            if cabecera_index is not None:
                with st.spinner("Generando archivo...", show_time=True):
                    df = pd.DataFrame(data[cabecera_index + 1:], columns=columnas)
                    df = df.reset_index(drop=True)
                    df = df[pd.isna(df["Severidad"]) & pd.isna(df["Probabilidad"]) & pd.isna(df["Ámbito"])]
                    st.write(f"Cantidad de registros sin completar: {len(df)}")
                    
                    if st.button("Comenzar proceso de análisis por LLM"):
                        filas_completadas = []
                        progress_text = "LLM trabajando..."
                        barra_progreso = st.progress(0, text=progress_text)
                        for i, row in df.iterrows():
                            st.divider()
                            with st.container(border=True):
                                st.write(row.get(columnas_descripcion)[0])
                                if pd.isna(row.get("Severidad")) and pd.isna(row.get("Probabilidad")) and pd.isna(row.get("Ámbito")):
                                    completado_raw = LLM_Consulta(client, system_prompt= system_instructions, descripcion=row.get(columnas_descripcion)[0])
                                    completado = json.loads(completado_raw)
                                    
                                    # Verificar si los valores son nulos antes de asignarlos
                                    if completado["SEVERIDAD"] is not None:
                                        df.at[i, "Severidad"] = completado["SEVERIDAD"]
                                    if completado["PROBABILIDAD"] is not None:
                                        df.at[i, "Probabilidad"] = completado["PROBABILIDAD"]
                                    if completado["AMBITO"] is not None:
                                        df.at[i, "Ámbito"] = completado["AMBITO"]
                                    
                                    # Añadir los campos adicionales solo si no son nulos
                                    if "PREGUNTAS" in completado and completado["PREGUNTAS"] is not None:
                                        if "Preguntas" not in df.columns:
                                            df["Preguntas"] = None
                                        # Si es una lista, unir elementos con retorno de carro
                                        if isinstance(completado["PREGUNTAS"], list):
                                            df.at[i, "Preguntas"] = "\n".join(completado["PREGUNTAS"])
                                        else:
                                            df.at[i, "Preguntas"] = completado["PREGUNTAS"]
                                        
                                    if "ANALISIS" in completado and completado["ANALISIS"] is not None:
                                        if "Análisis" not in df.columns:
                                            df["Análisis"] = None
                                        # Si es una lista, unir elementos con retorno de carro
                                        if isinstance(completado["ANALISIS"], list):
                                            df.at[i, "Análisis"] = "\n".join(completado["ANALISIS"])
                                        else:
                                            df.at[i, "Análisis"] = completado["ANALISIS"]
                                        
                                    if "RIESGOS" in completado and completado["RIESGOS"] is not None:
                                        if "Riesgos" not in df.columns:
                                            df["Riesgos"] = None
                                        # Si es una lista, unir elementos con retorno de carro
                                        if isinstance(completado["RIESGOS"], list):
                                            df.at[i, "Riesgos"] = "\n".join(completado["RIESGOS"])
                                        else:
                                            df.at[i, "Riesgos"] = completado["RIESGOS"]
                                        
                                    if "RECOMENDACIONES" in completado and completado["RECOMENDACIONES"] is not None:
                                        if "Recomendaciones" not in df.columns:
                                            df["Recomendaciones"] = None
                                        # Si es una lista, unir elementos con retorno de carro
                                        if isinstance(completado["RECOMENDACIONES"], list):
                                            df.at[i, "Recomendaciones"] = "\n".join(completado["RECOMENDACIONES"])
                                        else:
                                            df.at[i, "Recomendaciones"] = completado["RECOMENDACIONES"]

                                    # Mostrar solo si hay valores no nulos
                                    if all(val is not None for val in [completado.get("SEVERIDAD"), completado.get("PROBABILIDAD"), completado.get("AMBITO")]):
                                        severidad_badge = ":green-badge[" if completado['SEVERIDAD'] == 'Leve' else ":orange-badge[" if completado['SEVERIDAD'] == 'Grave' else ":red-badge[" if completado['SEVERIDAD'] == 'Muy Grave/Mortal' else ":blue-badge["
                                        probabilidad_badge = ":green-badge[" if completado['PROBABILIDAD'] == 'Improbable' else ":orange-badge[" if completado['PROBABILIDAD'] == 'Posible' else ":red-badge[" if completado['PROBABILIDAD'] == 'Probable' else ":blue-badge["
                                        ambito_badge = ":green-badge[" if completado['AMBITO'] == 'Puntual' else ":orange-badge[" if completado['AMBITO'] == 'Medio' else ":red-badge[" if completado['AMBITO'] == 'Extenso' else ":blue-badge["
                                        st.markdown(f"{severidad_badge}{completado['SEVERIDAD']}] {probabilidad_badge}{completado['PROBABILIDAD']}] {ambito_badge}{completado['AMBITO']}]", unsafe_allow_html=True)
                                        with st.expander("Conclusiones LLM 📝", expanded=False):
                                            col1, col2, col3, col4 = st.columns(4)
                                            with col1:
                                                st.subheader("Análisis", divider=True)
                                                st.write(completado["ANALISIS"])
                                            with col2:
                                                st.subheader("Preguntas", divider=True)
                                                st.write(completado["PREGUNTAS"])
                                            with col3:
                                                st.subheader("Riesgos", divider=True)
                                                st.write(completado["RIESGOS"])
                                            with col4:
                                                st.subheader("Recomendaciones", divider=True)
                                                st.write(completado["RECOMENDACIONES"]) 

                                    else:
                                        st.warning("No se pudieron obtener todas las puntuaciones para este registro.")

                                filas_completadas.append(i)
                                barra_progreso.progress((len(filas_completadas) / len(df)), text=progress_text)
                        st.success("Datos completados.")

                        if filas_completadas:
                            st.subheader("Registros completados:")
                            # Mostrar el dataframe con todas las columnas, no solo las básicas
                            cols_to_display = ["Severidad", "Probabilidad", "Ámbito", "Preguntas", "Análisis", "Riesgos", "Recomendaciones"]
                            # Asegurarse de mostrar también la columna de descripción
                            for col in columnas_descripcion:
                                if col in df.columns:
                                    cols_to_display.insert(0, col)
                            # Filtrar solo las columnas que existen en el dataframe
                            cols_to_display = [col for col in cols_to_display if col in df.columns]
                            st.dataframe(df.loc[filas_completadas, cols_to_display])
                        else:
                            st.info("No se encontraron registros para completar.")

                        # === Detectar índices de columnas ===
                        col_index_map = {col: idx for idx, col in enumerate(columnas)}

                        # Preparar las nuevas columnas antes de escribir
                        campos_adicionales = ["Preguntas", "Análisis", "Riesgos", "Recomendaciones"]
                        max_col = max(col_index_map.values()) + 1  # Índice de la última columna + 1
                        
                        # Añadir nuevas columnas al final si no existen
                        for campo in campos_adicionales:
                            if campo not in col_index_map:
                                columnas.append(campo)
                                col_index_map[campo] = max_col
                                # Establecer el encabezado en el Excel
                                ws.cell(row=cabecera_index + 1, column=max_col + 1).value = campo
                                max_col += 1

                        # === Escribir cambios en la hoja original ===
                        for idx in filas_completadas:
                            excel_row = cabecera_index + 2 + idx  # 1-based index en Excel
                            ws.cell(row=excel_row, column=col_index_map["Severidad"] + 1).value = df.at[idx, "Severidad"]
                            ws.cell(row=excel_row, column=col_index_map["Probabilidad"] + 1).value = df.at[idx, "Probabilidad"]
                            ws.cell(row=excel_row, column=col_index_map["Ámbito"] + 1).value = df.at[idx, "Ámbito"]
                            
                            # Escribir campos adicionales en el Excel
                            for campo in campos_adicionales:
                                if campo in df.columns and idx in df.index:
                                    valor = df.at[idx, campo]
                                    if valor is not None and not pd.isna(valor):
                                        # Si es un diccionario, convertirlo en un string con saltos de línea
                                        if isinstance(valor, dict):
                                            valor = '\n'.join([f"{k}: {v}" for k, v in valor.items()])
                                        
                                        # Asignar el valor a la celda
                                        celda = ws.cell(row=excel_row, column=col_index_map[campo] + 1)
                                        celda.value = valor

                                        # Aplicar ajuste de texto si contiene saltos de línea
                                        if "\n" in str(valor):
                                            celda.alignment = Alignment(wrapText=True)

                        # === Guardar el archivo con formato intacto ===
                        output = BytesIO()
                        wb.save(output)
                        output.seek(0)

                        st.download_button(
                            label="Descargar Excel Puntuaciones",
                            data=output,
                            file_name="Excel_puntuaciones.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
            else:
                st.error("No se encontró una cabecera con los campos requeridos: 'Severidad', 'Probabilidad' e 'Ámbito'.")
elif uploaded_file is not None and client is None:
    st.warning("Por favor, carga primero la configuración de Azure para continuar.")
elif uploaded_file is not None and not system_instructions:
    st.warning("Por favor, carga el archivo con el prompt del sistema antes de continuar.")

import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from pydantic import BaseModel
from openai import AzureOpenAI
import httpx
import json
from openpyxl.styles import Alignment


# Cargar las instrucciones del sistema con codificación UTF-8 explícita
with open("pages\system_prompt.md", "r", encoding="utf-8") as f:
    system_instructions = f.read()

# Configura tu cliente de Azure OpenAI
AZURE_CONFIG = {
    "deployment_name": "gpt-4o",
    "api_key": "cfe23f07a6f741ba85b1074adecc00b7",
    "azure_endpoint": "https://aisa-lab-factoria.openai.azure.com/",
    "api_version": "2025-01-01-preview"
} 

httpx_client = httpx.Client(http2=True, verify=False)

client = AzureOpenAI(
    api_key=AZURE_CONFIG["api_key"],  
    azure_endpoint=AZURE_CONFIG["azure_endpoint"],
    api_version=AZURE_CONFIG["api_version"],
    default_headers={"Content-Type": "application/json; charset=utf-8"},
    http_client=httpx_client
)
# Define tu modelo Pydantic
class puntuaciones(BaseModel):
    Severidad: str
    Probabilidad: str
    Ámbito: str

def LLM_Consulta(client, system_prompt = "", descripcion =""):
    # Realiza la solicitud al modelo desplegado
    completion = client.chat.completions.create(
        model="gpt-4o",  # Este es el nombre del *deployment*, no el modelo base. Usa el que configuraste en Azure.
        messages=[
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": descripcion},
        ],
        #response_format=puntuaciones,
        response_format={"type": "json_object"},
    )

      # Accede al contenido de la respuesta
    event = completion.choices[0].message.content
    #st.write(event)
    return(event)



st.title("Completar Excel con LLM")
apikey = st.text_input("APIKEY", type ="password")
uploaded_file = st.file_uploader("Sube tu archivo Excel", type=["xlsx"])

if uploaded_file is not None:
    in_memory_file = BytesIO(uploaded_file.read())

    # Cargar libro con openpyxl
    wb = load_workbook(filename=in_memory_file)
    hojas = wb.sheetnames
    hoja_seleccionada = st.selectbox("Selecciona una hoja", hojas)

    if hoja_seleccionada:
        ws = wb[hoja_seleccionada]
        data = list(ws.values)

        cabecera_index = None
        for i, fila in enumerate(data):
            if fila and "Severidad" in fila and "Probabilidad" in fila and "Ámbito" in fila:
                cabecera_index = i
                columnas = list(fila)

                # Buscar la(s) columna(s) que contienen "Descripción"
                columnas_descripcion = [col for col in columnas if "Descripción" in str(col)]
                
                if columnas_descripcion:
                    print(f"Columna(s) con 'Descripción' encontrada(s): {columnas_descripcion}")
                else:
                    print("No se encontró ninguna columna con 'Descripción'")



        if cabecera_index is not None:
            with st.spinner("Generando archivo...", show_time=True):
                df = pd.DataFrame(data[cabecera_index + 1:], columns=columnas)
                df = df.reset_index(drop=True)
                df = df[pd.isna(df["Severidad"]) & pd.isna(df["Probabilidad"]) & pd.isna(df["Ámbito"])]
                st.write(f"Cantidad de registros sin completar: {len(df)}")
                
                if st.button("Comenzar proceso de análisis por LLM"):
                    filas_completadas = []
                    progress_text = "LLM trabajando..."
                    barra_progreso = st.progress(0, text=progress_text)
                    for i, row in df.iterrows():
                        if pd.isna(row.get("Severidad")) and pd.isna(row.get("Probabilidad")) and pd.isna(row.get("Ámbito")):
                            completado_raw = LLM_Consulta(client, system_prompt= system_instructions, descripcion=row.get(columnas_descripcion)[0])
                            completado = json.loads(completado_raw)
                            
                            df.at[i, "Severidad"] = completado["SEVERIDAD"]
                            df.at[i, "Probabilidad"] = completado["PROBABILIDAD"]
                            df.at[i, "Ámbito"] = completado["AMBITO"]
                            
                            # Añadir los campos adicionales
                            if "PREGUNTAS" in completado:
                                if "Preguntas" not in df.columns:
                                    df["Preguntas"] = None
                                # Si es una lista, unir elementos con retorno de carro
                                if isinstance(completado["PREGUNTAS"], list):
                                    df.at[i, "Preguntas"] = "\n".join(completado["PREGUNTAS"])
                                else:
                                    df.at[i, "Preguntas"] = completado["PREGUNTAS"]
                                
                            if "ANALISIS" in completado:
                                if "Análisis" not in df.columns:
                                    df["Análisis"] = None
                                # Si es una lista, unir elementos con retorno de carro
                                if isinstance(completado["ANALISIS"], list):
                                    df.at[i, "Análisis"] = "\n".join(completado["ANALISIS"])
                                else:
                                    df.at[i, "Análisis"] = completado["ANALISIS"]
                                
                            if "RIESGOS" in completado:
                                if "Riesgos" not in df.columns:
                                    df["Riesgos"] = None
                                # Si es una lista, unir elementos con retorno de carro
                                if isinstance(completado["RIESGOS"], list):
                                    df.at[i, "Riesgos"] = "\n".join(completado["RIESGOS"])
                                else:
                                    df.at[i, "Riesgos"] = completado["RIESGOS"]
                                
                            if "RECOMENDACIONES" in completado:
                                if "Recomendaciones" not in df.columns:
                                    df["Recomendaciones"] = None
                                # Si es una lista, unir elementos con retorno de carro
                                if isinstance(completado["RECOMENDACIONES"], list):
                                    df.at[i, "Recomendaciones"] = "\n".join(completado["RECOMENDACIONES"])
                                else:
                                    df.at[i, "Recomendaciones"] = completado["RECOMENDACIONES"]
                            
                            filas_completadas.append(i)
                            barra_progreso.progress((len(filas_completadas) / len(df)), text=progress_text)
                    st.success("Datos completados.")

                    if filas_completadas:
                        st.subheader("Registros completados:")
                        st.dataframe(df.loc[filas_completadas])
                    else:
                        st.info("No se encontraron registros para completar.")

                    # === Detectar índices de columnas ===
                    col_index_map = {col: idx for idx, col in enumerate(columnas)}

                    # Preparar las nuevas columnas antes de escribir
                    campos_adicionales = ["Preguntas", "Análisis", "Riesgos", "Recomendaciones"]
                    max_col = max(col_index_map.values()) + 1  # Índice de la última columna + 1
                    
                    # Añadir nuevas columnas al final si no existen
                    for campo in campos_adicionales:
                        if campo not in col_index_map:
                            columnas.append(campo)
                            col_index_map[campo] = max_col
                            # Establecer el encabezado en el Excel
                            ws.cell(row=cabecera_index + 1, column=max_col + 1).value = campo
                            max_col += 1

                    # === Escribir cambios en la hoja original ===
                    for idx in filas_completadas:
                        excel_row = cabecera_index + 2 + idx  # 1-based index en Excel
                        ws.cell(row=excel_row, column=col_index_map["Severidad"] + 1).value = df.at[idx, "Severidad"]
                        ws.cell(row=excel_row, column=col_index_map["Probabilidad"] + 1).value = df.at[idx, "Probabilidad"]
                        ws.cell(row=excel_row, column=col_index_map["Ámbito"] + 1).value = df.at[idx, "Ámbito"]
                        
                        # Escribir campos adicionales en el Excel
                        for campo in campos_adicionales:
                            if campo in df.columns and idx in df.index:
                                valor = df.at[idx, campo]
                                if valor is not None and not pd.isna(valor):
                                    ws.cell(row=excel_row, column=col_index_map[campo] + 1).value = valor
                                    # Configurar la celda para que muestre correctamente el texto con saltos de línea
                                    cell = ws.cell(row=excel_row, column=col_index_map[campo] + 1)
                                    if "\n" in str(valor):
                                        cell.alignment = Alignment(wrapText=True)

                    # === Guardar el archivo con formato intacto ===
                    output = BytesIO()
                    wb.save(output)
                    output.seek(0)

                    st.download_button(
                        label="Descargar Excel Modificado",
                        data=output,
                        file_name="excel_modificado.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
        else:
            st.error("No se encontró una cabecera con los campos requeridos: 'Severidad', 'Probabilidad' e 'Ámbito'.")
